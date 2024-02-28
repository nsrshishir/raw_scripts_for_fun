from numpy import void
import openpyxl, os, shutil, requests
from pathlib import Path

folder_path = '/home/niaj/Desktop/idiya_images'
xlsx_file = Path('/home/niaj', 'Image_URL.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

def download_file(url, folder_path,local_filename):
    ext = url.split('.')[-1]
    path = os.path.join("{}/{}.{}".format(folder_path, local_filename,ext))
    print("{}\n{}".format(url,local_filename))
    if not os.path.exists(path):
        with requests.get(url, stream=True) as r:
            if r.ok:
                with open(path, 'wb') as f:
                    shutil.copyfileobj(r.raw, f)
            else:  # HTTP status code 4XX/5XX
                print("Download failed: {}".format(local_filename))
    return void

col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
   
    
print(col_names)

data={}
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    print(i)    
    if i == 0:
       continue
        
    else:
        download_file(row[0], folder_path, row[1])



